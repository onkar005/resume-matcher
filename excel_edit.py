from openpyxl import load_workbook

def edit_excel(file_name,sheet_name,cell_name,edit_value):
	wb = load_workbook(file_name)
	ws = wb[sheet_name]
	ws[cell_name] = edit_value
	wb.save(file_name)

def read_excel(file_name,sheet_name,cell_name):
	wb = load_workbook(file_name)
	ws = wb[sheet_name]
	cell_value = ws[cell_name]
	return cell_value.value


